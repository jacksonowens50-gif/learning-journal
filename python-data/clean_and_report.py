"""
clean_and_report.py — Day 7

Reads a messy sales CSV, cleans it, logs every change made, and writes a
formatted multi-sheet Excel workbook (Summary / Detail / Cleaning Log).

Usage:
    python clean_and_report.py messy_sales.csv
"""

import sys
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# Every cleaning function appends a record here. This is the audit trail.
CLEANING_LOG = []

# Column name -> Excel number format. Applied wherever the name appears.
FORMATS = {
    'sales': '#,##0.00',
    'profit': '#,##0.00',
    'discount': '0%',
    'quantity': '#,##0',
    'orders': '#,##0',
    'margin': '0.0%',
    'postal_code': '0',
    'order_date': 'yyyy-mm-dd',
    'ship_date': 'yyyy-mm-dd',
}

def log(step, issue, rows_affected, action):
    """Record one cleaning decision so it can be reported to the client."""
    CLEANING_LOG.append({
        'Step': step,
        'Issue': issue,
        'Rows Affected': rows_affected,
        'Action Taken': action,
    })


def load_raw(path):
    """Read the CSV, skipping export junk above the header. Returns a DataFrame."""
    df = pd.read_csv(path, encoding='latin-1', skiprows=4)
    log('Load', f'Read source file: {path}', len(df), 'skiprows=4 (export metadata), encoding=latin-1')
    return df


def clean_headers(df):
    """Normalize column names: strip, lowercase, underscores."""
    before = list(df.columns)
    df.columns = (df.columns
                  .str.strip()
                  .str.lower()
                  .str.replace(r'[^0-9a-z]+', '_', regex=True)
                  .str.strip('_'))
    df = df.rename(columns={'customer': 'customer_id'})
    changed = sum(1 for b, a in zip(before, df.columns) if b != a)
    log('Headers', 'Column names had mixed case, trailing whitespace, and symbols (#, /)',
        changed, 'Normalized all to lowercase_with_underscores; customer -> customer_id')
    return df


def fix_types(df):
    """Coerce sales to numeric and date columns to datetime. Log failures."""
    already_null = df['sales'].isna().sum()
    stripped = df['sales'].astype(str).str.replace(r'[$,]', '', regex=True)
    df['sales'] = pd.to_numeric(stripped, errors='coerce')
    coerced = df['sales'].isna().sum() - already_null
    log('Types',
        "sales stored as text with '$' prefixes, thousands separators, and '-' placeholders",
        coerced,
        f"Stripped $ and commas; {coerced} non-numeric values coerced to null (NOT zero)")

    for col in ['order_date', 'ship_date']:
        before_null = df[col].isna().sum()
        df[col] = pd.to_datetime(df[col], errors='coerce')
        failed = df[col].isna().sum() - before_null
        log('Types', f'{col} stored as text', failed,
            f'Parsed to datetime; {failed} values unparseable')

    return df


def handle_missing(df):
    """Apply the documented missing-value decisions. Impute nothing silently."""
    missing_postal = df['postal_code'].isna().sum()
    log('Missing values', 'postal_code blank', missing_postal,
        'Left as null — geographic label, not a measure. Imputing would fabricate data.')

    missing_sales = df['sales'].isna().sum()
    log('Missing values', 'sales blank or unreadable', missing_sales,
        'Left as null, excluded from aggregates via pandas skipna. NOT filled with 0 — '
        'zero is a real value and would understate averages and margin.')

    return df


def remove_duplicates(df):
    """Drop exact duplicate rows. Flag ambiguous business-key duplicates."""
    exact = df.duplicated().sum()
    df = df.drop_duplicates(keep='first').reset_index(drop=True)
    log('Duplicates', 'Exact duplicate rows — all 21 fields identical', exact,
        'Dropped, kept first occurrence. Defensible without client input: an identical '
        'row cannot represent two distinct transactions.')

    key = ['order_id', 'product_id']
    ambiguous = df.duplicated(subset=key, keep=False).sum()
    log('Duplicates',
        'Rows sharing order_id + product_id but differing in quantity and sales',
        ambiguous,
        'RETAINED — not dropped. Unit prices are consistent within each pair, suggesting '
        'legitimate split line items. OPEN: requires client confirmation.')

    non_id = [c for c in df.columns if c != 'row_id']
    near_exact = df.duplicated(subset=non_id).sum()
    log('Duplicates', 'Rows identical on every field except row_id', near_exact,
        'RETAINED and escalated — strong candidate for true double-entry in the source '
        'system. OPEN: requires client confirmation before removal.')

    return df

def build_summary(df):
    """Return the exec-view summary tables: totals, breakdown, flags."""
    total_sales = df['sales'].sum()
    total_profit = df['profit'].sum()

    kpis = pd.DataFrame(
        [
            ('Rows in report', len(df)),
            ('Unique orders', df['order_id'].nunique()),
            ('Date range', f"{df['order_date'].min():%Y-%m-%d} to {df['order_date'].max():%Y-%m-%d}"),
            ('Total sales', total_sales),
            ('Total profit', total_profit),
            ('Margin %', total_profit / total_sales),
            ('Sales values excluded as null', int(df['sales'].isna().sum())),
        ],
        columns=['Metric', 'Value'],
    )

    breakdown = df.groupby('category', as_index=False).agg(
        sales=('sales', 'sum'),
        profit=('profit', 'sum'),
        orders=('order_id', 'nunique'),
    )
    breakdown['margin'] = breakdown['profit'] / breakdown['sales']
    breakdown = breakdown.sort_values('sales', ascending=False)

    return {'kpis': kpis, 'breakdown': breakdown}


def style_sheet(ws, header_row=1):
    """Bold the header row and size every column to its longest value."""
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
    for i, col in enumerate(ws.iter_cols(), start=1):
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 45)


def apply_formats(ws, header_row=1):
    """Apply FORMATS to any column whose header matches, below header_row."""
    headers = {c.value: c.column for c in ws[header_row]}
    for name, fmt in FORMATS.items():
        if name not in headers:
            continue
        col = headers[name]
        for r in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = fmt


def write_excel(df, summary, out_path):
    """Write Summary / Detail / Cleaning Log sheets with formatting."""
    log_df = pd.DataFrame(CLEANING_LOG)
    kpi_rows = len(summary['kpis'])
    breakdown_start = kpi_rows + 3          # blank rows between the two tables

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        summary['kpis'].to_excel(writer, sheet_name='Summary', index=False)
        summary['breakdown'].to_excel(writer, sheet_name='Summary', index=False,
                                      startrow=breakdown_start)
        df.to_excel(writer, sheet_name='Detail', index=False)
        log_df.to_excel(writer, sheet_name='Cleaning Log', index=False)

        # --- Summary sheet ---
        ws = writer.sheets['Summary']
        style_sheet(ws, header_row=1)
        for cell in ws[breakdown_start + 1]:
            cell.font = Font(bold=True)
        apply_formats(ws, header_row=breakdown_start + 1)

        # KPI values are mixed types, so format them cell by cell.
        kpi_formats = {
            'Rows in report': '#,##0',
            'Unique orders': '#,##0',
            'Total sales': '$#,##0.00',
            'Total profit': '$#,##0.00',
            'Margin %': '0.0%',
            'Sales values excluded as null': '#,##0',
        }
        for r in range(2, kpi_rows + 2):
            metric = ws.cell(row=r, column=1).value
            if metric in kpi_formats:
                ws.cell(row=r, column=2).number_format = kpi_formats[metric]

        # --- Detail sheet ---
        detail = writer.sheets['Detail']
        style_sheet(detail)
        apply_formats(detail)
        detail.freeze_panes = 'A2'
        detail.auto_filter.ref = detail.dimensions

        # --- Cleaning Log sheet ---
        logsheet = writer.sheets['Cleaning Log']
        style_sheet(logsheet)
        logsheet.freeze_panes = 'A2'
        for row in logsheet.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        logsheet.column_dimensions['D'].width = 70


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'messy_sales.csv'

    df = load_raw(path)
    df = clean_headers(df)
    df = fix_types(df)
    df = handle_missing(df)
    df = remove_duplicates(df)

    summary = build_summary(df)
    write_excel(df, summary, 'sales_report.xlsx')

    print('done')


if __name__ == '__main__':
    main()

